import datetime

from django.db import models
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
import os


class Customer(models.Model):
    first = models.CharField(max_length=100)
    last = models.CharField(max_length=100)
    phone = models.CharField(max_length=100)
    email = models.CharField(max_length=100)
    street = models.CharField(max_length=100)
    apt_suite_other = models.CharField(max_length=100)
    city = models.CharField(max_length=100)
    state = models.CharField(max_length=100)
    zip_code = models.CharField(max_length=15)

    def full_address(self):
        return f'{self.street} {self.apt_suite_other} {self.city}, {self.state}, {self.zip_code}'

    def __str__(self):
        return f'{self.first} {self.last}'

    def full_name(self):
        return f'{self.first} {self.last}'


class RentalOrder(models.Model):
    invoice = models.CharField(max_length=6)
    date = models.DateField('date')
    customer = models.ForeignKey(Customer, on_delete=models.CASCADE)
    lg_boxes = models.IntegerField(default=0)
    xl_boxes = models.IntegerField(default=0)
    lg_dollies = models.IntegerField(default=0)
    labels = models.IntegerField(default=0)
    zip_ties = models.IntegerField(default=0)
    bins = models.IntegerField(default=0)
    rental_period = models.CharField(max_length=100)
    delivery_date = models.DateField('delivery date')
    delivery_street = models.CharField(max_length=100)
    delivery_apt_suite_other = models.CharField(max_length=100)
    delivery_city = models.CharField(max_length=100)
    delivery_state = models.CharField(max_length=100)
    delivery_zip_code = models.CharField(max_length=15)
    pickup_date = models.DateField('pickup date')
    total_price = models.FloatField(default=0)
    pickup_address = models.CharField(max_length=100)

    def __str__(self):
        return f''' 
        {self.invoice}*** {self.customer} {self.date}
        '''

    def details(self):
        return f''' 
        Invoice #: {self.invoice}\n
        Customer: {self.customer}\n
        Date: {self.date}\n
        Lg Boxes: {self.lg_boxes}\n
        Xl Boxes: {self.xl_boxes}\n
        Lg Dollies: {self.lg_dollies}\n
        Labels: {self.labels}\n
        Zip-Ties: {self.zip_ties}\n 
        Bins: {self.bins}\n
        Rental Period: {self.rental_period}
        Delivery Date: {self.delivery_date}
        Pickup Date: {self.pickup_date}
        '''

    def recent_order(self):
        now = timezone.now()
        last_week = now - datetime.timedelta(days=7)
        return  last_week.date() <= self.date <= now.date()

    def calculate_total(self):
        pass

    def was_delivered(self):
        now = timezone.now().date()
        return self.delivery_date <= now

    def was_picked_up(self):
        now = timezone.now().date()
        return self.pickup_date <= now

    def upcoming(self):
        now = timezone.now()
        next_week = now + datetime.timedelta(days=7)
        if now.date() <= self.delivery_date <= next_week.date():
            return True

        elif now.date() <= self.pickup_date <= next_week.date():
            return True

    def write_workbook(self):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{self.customer.full_name()}{self.invoice}.xlsx"'

        path = os.path.dirname(__file__)
        file = os.path.join(path, 'rental_order.xlsx')
        wb = load_workbook(file)
        ws = wb.active

        ws['G3'] = self.date
        ws['G4'] = self.invoice
        ws['C7'] = self.customer.full_name()
        ws['C8'] = self.customer.street
        ws['C9'] = f'{self.customer.city}, {self.customer.state} {self.customer.zip_code}'
        ws['C10'] = self.customer.phone
        ws['C11'] = self.customer.email
        ws['F7'] = self.delivery_street
        ws['F8'] = f'{self.delivery_city}, {self.delivery_state} {self.delivery_zip_code}'
        ws['F11'] = self.pickup_address
        ws['F16'] = self.delivery_date
        ws['G16'] = self.pickup_date
        ws['B18'] = self.rental_period
        ws['G18'] = self.was_delivered()
        ws['G34'] = self.total_price
        ws['B21'] = self.lg_boxes
        ws['B22'] = self.xl_boxes
        ws['B23'] = self.lg_dollies
        ws['B24'] = self.xl_dollies
        ws['B25'] = self.wardrobes
        ws['B26'] = self.labels
        ws['B27'] = self.zip_ties
        ws['B28'] = self.bins
        ws['C21'] = 'LG Boxes'
        ws['C22'] = 'XL Boxes'
        ws['C23'] = 'LG Dollies'
        ws['C24'] = 'XL Dollies'
        ws['C25'] = 'Wardrobes'
        ws['C26'] = 'Labels'
        ws['C27'] = 'Zip Ties'
        ws['C28'] = 'Bins'

        wb.save(response)
        return response

    def delivery_address(self):
        if self.delivery_apt_suite_other:
            return f'{self.delivery_street}, {self.delivery_apt_suite_other},' \
                   f' {self.delivery_city}, {self.delivery_state} {self.delivery_zip_code}'
        else:
            return f'{self.delivery_street}, {self.delivery_city}, {self.delivery_state} {self.delivery_zip_code}'


    # def pickup_address(self):
    #     if self.delivery_apt_suite_other:
    #         return f'{self.pickup_street}, {self.pickup_apt_suite_other},' \
    #                f' {self.pickup_city} {self.pickup_state} {self.pickup_zip_code}'
    #     else:
    #         return f'{self.pickup_street}, {self.pickup_city}, {self.pickup_state} {self.pickup_zip_code}'

class Location(models.Model):
    zip_code = models.CharField(max_length=16)
    name = models.CharField(max_length=255)
    description = models.TextField

    def __str__(self):
        return f'{self.name}'

class ItemType(models.Model):
    type_name = models.CharField(max_length=64)

    def __str__(self):
        return f'{self.type_name}'

class Item(models.Model):
    item_name = models.CharField(max_length=100)
    item_type_id = models.ForeignKey(ItemType, on_delete=models.CASCADE)
    location_id = models.ForeignKey(Location, on_delete=models.CASCADE)
    item_location = models.CharField(max_length=100)
    description = models.CharField(max_length=100)
    price_per_unit = models.DecimalField(max_digits=10, decimal_places=2)
    unit_id = models.CharField(max_length=100)
    available = models.CharField(max_length=100)


