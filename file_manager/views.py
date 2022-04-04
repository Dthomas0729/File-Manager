from __future__ import print_function

import os
import os.path

from decouple import config
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.core.mail import send_mail
from django.shortcuts import render, get_object_or_404, redirect
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import load_workbook
from woocommerce import API
from datetime import datetime, timedelta


from .forms import CreateUserForm, CustomerForm, RentalOrderForm
from .models import Customer, RentalOrder

def handle_error_email(order):

    send_mail(
    'Order Entered Incorrectly, PLEASE FIX!',
    'Here is the message.',
    'slicctech@gmail.com', 
    ['d.thomas0729@gmail.com'],
    fail_silently=False,
)

def new_order_email(order):
    send_mail(
    'New Order Added!',
    f''' Here is the order:
    {order.details()}
    http://127.0.0.1:8000/{order.invoice}
    ''',
    'slicctech@gmail.com',
    ['d.thomas0729@gmail.com'],
    fail_silently=False,
)

# This will establish the WooCommerce API to make calls
wcapi = API(
    url=config('WCAPI_URL'),
    consumer_key=config('WCAPI_CONSUMER_KEY'),
    consumer_secret=config('WCAPI_CONSUMER_SECRET'),
    version="wc/v3"
)

data = wcapi.get('orders').json()

def update_order_db():
    current_order = data[0]
    f_name = current_order['billing']['first_name']
    l_name = current_order['billing']['last_name']
    phone = current_order['billing']['phone']
    email = current_order['billing']['email']
    street = current_order['billing']['address_1']
    apt_suite_other = current_order['billing']['address_2']
    city = current_order['billing']['city']
    state = current_order['billing']['state']
    zip_code = current_order['billing']['postcode']

    last_customer = Customer(
        first=f_name,
        last=l_name,
        phone=phone,
        email=email,
        street=street,
        apt_suite_other=apt_suite_other,
        city=city,
        state=state,
        zip_code=zip_code
    )

    try:
        last_customer = Customer.objects.get(first=last_customer.first, last=last_customer.last)
    except Customer.DoesNotExist:
        last_customer.save()

    invoice = current_order['id']
    date = datetime.strptime(current_order['date_created'], '%Y-%m-%dT%H:%M:%S').date()
    delivery_street = current_order['shipping']['address_1']
    delivery_apt_suite_other = current_order['shipping']['address_2']
    delivery_city = current_order['shipping']['city']
    delivery_state = current_order['shipping']['state']
    delivery_zip_code = current_order['shipping']['postcode']
    pickup_address = current_order['meta_data'][5]['value']
    total_price = current_order['total']

    delivery_datetime = current_order['meta_data'][0]['value'] + current_order['meta_data'][1]['value']
    try:
        delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%Y-%m-%d')
    except ValueError:
        try:
            delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%m/%d/%Y')
        except ValueError:
            try:
                delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%B %d, %Y')
            except ValueError:
                try:
                    delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%B %dth %Y')
                except ValueError:
                    try:
                        delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%B %st %Y')
                    except ValueError:
                        try:
                            delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%B %dnd %Y')
                        except ValueError:
                            delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%B %drd %Y')

    try:
        delivery_datetime = datetime.strptime(delivery_datetime, '%Y-%m-%d%H:%M')
    except ValueError:
        try:
            delivery_datetime = datetime.strptime(delivery_datetime, '%m/%d/%Y%H:%M %p')
        except ValueError:
            try:
                delivery_datetime = datetime.strptime(delivery_datetime, '%B %dth %Y%H%p')
            except ValueError:
                try:
                    delivery_datetime = datetime.strptime(delivery_datetime, '%B %d, %Y%H%p')
                except ValueError:
                    delivery_datetime = None

    delivery_date = delivery_date.date()

    # THIS LOCATES THE RENTAL PERIOD AND CREATES PICKUP DATETIME OBJECT
    try:
        rental_period = int(current_order['line_items'][0]['meta_data'][0]['value'][0])
        if rental_period == 1:
            pickup_date = delivery_date + timedelta(days=7)
            rental_period = '1 Week'
        elif rental_period == 2:
            pickup_date = delivery_date + timedelta(days=14)
            rental_period = '2 Weeks'
        elif rental_period == 3:
            pickup_date = delivery_date + timedelta(days=21)
            rental_period = '3 Weeks'
        elif rental_period == 4:
            pickup_date = delivery_date + timedelta(days=28)
            rental_period = '4 Weeks'
    except IndexError:
        handle_error_email()
        rental_period = '1 Week'
        pickup_date = delivery_date + timedelta(days=1)
    except KeyError:
        handle_error_email()
        rental_period = '1 Week'
        pickup_date = delivery_date + timedelta(days=1)

    for x in range(len(current_order['line_items'])):
        if current_order['line_items'][x]['product_id'] == 1270:
            lg_boxes = 75
            xl_boxes = 10
            lg_dollies = 4
            labels = 85
            zip_ties = 85
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1515:
            lg_boxes = 55
            xl_boxes = 10
            lg_dollies = 3
            labels = 65
            zip_ties = 65
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1510:
            lg_boxes = 40
            xl_boxes = 5
            lg_dollies = 2
            labels = 45
            zip_ties = 45
            bins = 0
        elif current_order['line_items'][x]['product_id'] == 1505:
            lg_boxes = 22
            xl_boxes = 3
            lg_dollies = 2
            labels = 25
            zip_ties = 25
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1545:
            lg_boxes = 1
            xl_boxes = 0
            lg_dollies = 0
            xl_dollies = 0
            labels = 0
            zip_ties = 0
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1291:
            lg_boxes = 0
            xl_boxes = 0
            lg_dollies = 0
            xl_dollies = 0
            labels = 0
            zip_ties = 0
            bins = current_order['line_items'][x]['quantity']
            break
        else:
            lg_boxes = 0
            xl_boxes = 0
            lg_dollies = 0
            xl_dollies = 0
            labels = 0
            zip_ties = 0
            bins = 0

    last_order = RentalOrder(
        invoice=invoice,
        date=date,
        customer=last_customer,
        lg_boxes=lg_boxes,
        xl_boxes=xl_boxes,
        lg_dollies=lg_dollies,
        labels=labels,
        zip_ties=zip_ties,
        bins=bins,
        handtrucks=0,
        delivery_date=delivery_date,
        delivery_street=delivery_street,
        delivery_apt_suite_other=delivery_apt_suite_other,
        delivery_city=delivery_city,
        delivery_state=delivery_state,
        delivery_zip_code=delivery_zip_code,
        rental_period=rental_period,
        pickup_date=pickup_date,
        total_price=total_price,
        pickup_address=pickup_address,
    )

    try:
        RentalOrder.objects.get(invoice=last_order.invoice)
        return [last_order, last_customer]

    except RentalOrder.DoesNotExist:
        last_order.save()
        new_order_email(last_order)
        return [last_order, last_customer]


def post_events(delivery_event, pickup_event):

    SCOPES = ['https://www.googleapis.com/auth/calendar']

    service_account_file = 'taggabox-web-app-d02e1e73d5cf.json'

    credentials = service_account.Credentials.from_service_account_file(
        service_account_file, scopes=SCOPES)
    credentials = credentials.with_subject('taggabox-order-update@flowing-blade-284003.iam.gserviceaccount.com')

    service = build('calendar', 'v3', credentials=credentials)

# Call the Calendar API
    service.events().insert(calendarId='primary', body=delivery_event).execute()
    service.events().insert(calendarId='primary', body=pickup_event).execute()


def create_delivery_event(last_order, customer):

    start_time = last_order.delivery_date
    start_time = start_time.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    end_time = last_order.delivery_date + timedelta(hours=1)
    end_time = end_time.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    delivery_event = {
        'summary': f'{customer.first} {customer.last} Box Delivery',
        'location': f'{last_order.delivery_street}, {last_order.delivery_city}, {last_order.delivery_state} {last_order.delivery_zip_code}',
        'description': f'''{customer.first} {customer.last}
phone: {customer.phone}
address: {last_order.delivery_street}, {last_order.delivery_city}, {last_order.delivery_state} {last_order.delivery_zip_code}
email: {customer.email}
{last_order.lg_boxes} Lg Boxes
{last_order.xl_boxes} Xl Boxes
{last_order.lg_dollies} Dollies
{last_order.labels} Labels & Zip Ties
{last_order.bins} Bins''',

        'start': {
            'dateTime': f'{start_time}',
            'timeZone': 'America/New_York',
        },
        'end': {
            'dateTime': f'{end_time}',
            'timeZone': 'America/New_York',
        },
        'reminders': {
            'useDefault': False,
            'overrides': [
                {'method': 'email', 'minutes': 24 * 60},
                {'method': 'popup', 'minutes': 10},
            ],
        },
    }
    return delivery_event


def create_pickup_event(last_order, customer):
    start_time = last_order.pickup_date
    start_time = start_time.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    end_time = last_order.pickup_date + timedelta(hours=1)
    end_time = end_time.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    pickup_event = {
        'summary': f'{customer.first} {customer.last} Box Pick Up',
        'location': f'{last_order.pickup_address}',
        'description': f'''{customer.first} {customer.last}
phone: {customer.phone}
Address {last_order.pickup_address}
email: {customer.email}
{last_order.lg_boxes} Lg Boxes                                                                      
{last_order.xl_boxes} Xl Boxes
{last_order.lg_dollies} Dollies
{last_order.labels} Labels & Zip Ties
{last_order.bins} Bins''',

        'start': {
            'dateTime': f'{start_time}',
            'timeZone': 'America/New_York',
        },
        'end': {
            'dateTime': f'{end_time}',
            'timeZone': 'America/New_York',
        },
        'reminders': {
            'useDefault': False,
            'overrides': [
                {'method': 'email', 'minutes': 24 * 60},
                {'method': 'popup', 'minutes': 10},
            ],
        },
    }
    return pickup_event



def new_order_webhook(request):
    order, customer = update_order_db()

    delivery = create_delivery_event(order, customer)
    pickup = create_pickup_event(order, customer)

    # post_events(delivery, pickup)
    return HttpResponse('Hello, world. This is the webhook response.')

# THIS ALLOWS FOR THE USER TO DOWNLOAD AN EXCEL FILE CONTAINING RENTAL ORDER & CUSTOMER INFORMATION
def export_order_file(request, invoice):

    # RETRIEVE ORDER DETAILS FROM DB
    details = get_object_or_404(RentalOrder, invoice=invoice)

    # NOT EXACTLY SURE HOW THIS CODE WORKS
    response = HttpResponse(content_type='application/ms-excel')

    # CREATES NEW FILENAME FOR WORKBOOK TO BE SAVED AS
    response['Content-Disposition'] = f'attachment; filename="{details.customer.full_name()}{details.invoice}.xlsx"'

    path = os.path.dirname(__file__)
    file = os.path.join(path, 'rental_order.xlsx')

    # LOAD TEMPLATE WORKBOOK FROM FILE
    wb = load_workbook(file)

    # LOAD WORKSHEET FROM WORKBOOK.ACTIVE AND ASSIGN ALL NECESSARY INFO INTO WS CELLS
    ws = wb.active
    ws['F3'] = details.date
    ws['F4'] = details.invoice
    ws['C7'] = details.customer.full_name()
    ws['C8'] = details.customer.street
    ws['C9'] = f'{details.customer.city}, {details.customer.state} {details.customer.zip_code}'
    ws['C10'] = details.customer.phone
    ws['C11'] = details.customer.email
    ws['F7'] = details.delivery_street
    ws['F8'] = f'{details.delivery_city}, {details.delivery_state} {details.delivery_zip_code}'
    ws['F11'] = details.pickup_address
    ws['F16'] = details.delivery_date
    ws['F18'] = details.pickup_date
    ws['B18'] = details.rental_period
    # ws['G18'] = details.was_delivered()
    # ws['G34'] = details.total_price
    ws['B21'] = details.lg_boxes
    ws['B22'] = details.xl_boxes
    ws['B23'] = details.lg_dollies
    # ws['B24'] = details.xl_dollies
    ws['B25'] = details.labels
    ws['B26'] = details.zip_ties
    ws['B27'] = details.bins
    ws['B28'] = details.handtrucks

    # SAVE WORKBOOK INTO NEW FILENAME SAVED INSIDE RESPONSE VARIABLE
    wb.save(response)
    return response


color_theme = '#065821'


# VIEW PAGES ARE BELOW
@login_required(login_url='/login')
def home(request):

    search = request.POST.get('search')

    # THIS CREATES LIST OBJECT OF RECENT & UPCOMING ORDERS ORDERED BY DATE
    last_orders = RentalOrder.objects.all()
    recent_orders = []
    upcoming_orders = []
    for order in last_orders:
        if order.recent_order():
            recent_orders.append(order)

    for order in last_orders:
        if order.upcoming():
            upcoming_orders.append(order)

    # CREATE A DICT OBJ CONTEXT FOR NEW LIST
    context = {
        'recent_orders': recent_orders,
        'upcoming_orders': upcoming_orders,
        'search': search,
    }

    return render(request, 'file_manager/home.html', context)


@login_required(login_url='/login')
def new_search(request):
    search = request.POST.get('search').capitalize()
    try:
        c_results = Customer.objects.get(first=search)
    except Customer.DoesNotExist:
        results = None
    else:
        results = RentalOrder.objects.filter(customer=c_results)

    context = {
        'search': search,
        'orders': results
    }

    return render(request, 'file_manager/new_search.html', context)


@login_required(login_url='/login')
def order_details(request, invoice):
    details = get_object_or_404(RentalOrder, invoice=invoice)

    context = {
        'details': details,
    }
    return render(request, 'file_manager/orders/order_details.html', context)


# CREATE DELIVERY AND PICKUP EVENTS TO POST TO GOOGLE CALENDAR




@login_required(login_url='/login')
def customer_details(request):
    search = request.POST.get('search')
    context = {
        'search': search,
    }

    return render(request, 'file_manager/customer_details.html', context)


# HANDLES USER INPUT FOR CUSTOMER FORM TO SAVE NEW CUSTOMER INTO DB
@login_required(login_url='/login')
def customer_form(request):
    form = CustomerForm()

    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()

    context = {'form': form,
               }
    return render(request, 'file_manager/customer_form.html', context)


# HANDLES USER INPUT FOR RENTAL ORDER FORM TO SAVE NEW ORDER INTO DB
@login_required(login_url='/login')
def rental_order_form(request):
    form = RentalOrderForm()

    if request.method == 'POST':
        form = RentalOrderForm(request.POST)
        if form.is_valid():
            form.save()

    context = {'form': form}
    return render(request, 'file_manager/orders/rental_order_form.html', context)


# HANDLES USER INPUT FOR UPDATING EXISTING ORDERS AND SAVES NEW INFO INTO DB
@login_required(login_url='/login')
def update_order_form(request, pk):

    order = RentalOrder.objects.get(id=pk)
    form = RentalOrderForm(instance=order)

    if request.method == 'POST':
        form = RentalOrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('/orders')
    context = {'form': form}
    return render(request, 'file_manager/orders/update_order.html', context)


@login_required(login_url='/login')
def delete_order(request, pk):
    order = RentalOrder.objects.get(id=pk)

    if request.method == 'POST':
        order.delete()
        return redirect('/orders')

    context = {'order': order}
    return render(request, 'file_manager/orders/delete.html', context)


# THIS VIEW LISTS MOST RECENT RENTAL ORDERS
@login_required(login_url='/login')
def display_orders(request):
    # data = wcapi.get('orders?page=4').json()

    # for order in range(len(data)):
    #     views.update_order_db(order)

    # RETRIEVE RENTAL ORDER LIST FROM DJANGO DB ORDERED BY MOST RECENT INVOICE
    rental_orders = RentalOrder.objects.all().order_by('-invoice')

    # PAGINATOR CLASS WILL LIST ORDERS 10 PER PAGE
    p = Paginator(rental_orders, 10)
    page_num = request.GET.get('page')
    page = p.get_page(page_num)

    context = {
        'page_obj': p,
        'page': page
    }
    return render(request, 'file_manager/orders/all_orders.html', context)


@login_required(login_url='/login')
def inventory(request):
    return render(request, 'file_manager/inventory/inventory.html')


def login_page(request):
    if request.user.is_authenticated:
        return redirect('home')
    else:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')

            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                messages.info(request, 'Username OR Password is incorrect')

        return render(request, 'file_manager/registration/login.html')



@login_required(login_url='/login')
def register_page(request):
    if request.user.is_authenticated:
        return redirect('home')
    else:
        form = CreateUserForm()

        if request.method == 'POST':
            form = CreateUserForm(request.POST)
            if form.is_valid():
                form.save()
                user = form.cleaned_data.get('username')
                messages.success(request, 'Account was created for ' + user)

                return redirect('/login')

        context = {
            'form': form,
        }
        return render(request, 'file_manager/registration/register.html', context)


def logout_user(request):
    logout(request)

    return redirect('/login')

